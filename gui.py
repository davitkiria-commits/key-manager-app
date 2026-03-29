from __future__ import annotations

import shutil
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Callable

import requests
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QIntValidator
from PySide6.QtWidgets import (
    QComboBox,
    QDateEdit,
    QDialog,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from key_manager import KeyManager, KeyManagerError, Person


UPDATE_CONFIG = {
    "github_username": "davitkiria-commits",
    "github_repo": "key-manager-app",
    "github_branch": "main",
}


def get_update_urls() -> dict[str, str]:
    username = UPDATE_CONFIG["github_username"]
    repo = UPDATE_CONFIG["github_repo"]
    branch = UPDATE_CONFIG["github_branch"]
    base_raw_url = f"https://raw.githubusercontent.com/{username}/{repo}/{branch}"
    return {
        "version_txt_url": f"{base_raw_url}/version.txt",
        "main_zip_url": f"https://github.com/{username}/{repo}/archive/refs/heads/{branch}.zip",
    }


class AddApartmentDialog(QDialog):
    def __init__(
        self,
        manager: KeyManager,
        apartment_id: int | None = None,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.manager = manager
        self.apartment_id = apartment_id
        self.apartment = self.manager.get_apartment(apartment_id) if apartment_id is not None else None

        is_edit = self.apartment is not None
        self.setWindowTitle("Редактировать квартиру" if is_edit else "Добавить квартиру")

        self.building_edit = QLineEdit(self.apartment.building if self.apartment else "")
        self.floor_edit = QLineEdit(str(self.apartment.floor) if self.apartment else "")
        self.floor_edit.setValidator(QIntValidator(-100, 300, self))
        self.apartment_edit = QLineEdit(self.apartment.apartment_number if self.apartment else "")
        self.total_keys_spin = QSpinBox()
        self.total_keys_spin.setRange(0, 1000)
        if self.apartment:
            self.total_keys_spin.setValue(self.apartment.total_keys)
        else:
            self.total_keys_spin.setValue(1)

        form = QFormLayout()
        form.addRow("Корпус:", self.building_edit)
        form.addRow("Этаж:", self.floor_edit)
        form.addRow("Квартира:", self.apartment_edit)
        form.addRow("Всего ключей:", self.total_keys_spin)

        save_btn = QPushButton("Сохранить" if is_edit else "Добавить")
        save_btn.clicked.connect(self._save)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(save_btn)

    def _save(self) -> None:
        try:
            building = self.building_edit.text().strip()
            floor_text = self.floor_edit.text().strip()
            apt = self.apartment_edit.text().strip()
            total = self.total_keys_spin.value()

            if not building or not floor_text or not apt:
                raise KeyManagerError("Заполните все поля.")

            floor = int(floor_text)
            if self.apartment:
                self.manager.update_apartment(
                    apartment_id=self.apartment.apartment_id,
                    building=building,
                    floor=floor,
                    apartment_number=apt,
                    total_keys=total,
                )
            else:
                self.manager.add_apartment(building=building, floor=floor, apartment_number=apt, total_keys=total)
            self.accept()
        except (ValueError, KeyManagerError) as e:
            QMessageBox.warning(self, "Ошибка", str(e))


class PersonEditDialog(QDialog):
    def __init__(self, manager: KeyManager, person: Person | None = None, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.person = person

        self.setWindowTitle("Редактировать получателя" if person else "Добавить получателя")

        self.name_edit = QLineEdit(person.full_name if person else "")
        self.role_edit = QLineEdit(person.role if person else "")

        form = QFormLayout()
        form.addRow("ФИО:", self.name_edit)
        form.addRow("Роль:", self.role_edit)

        save_btn = QPushButton("Сохранить")
        save_btn.clicked.connect(self._save)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(save_btn)

    def _save(self) -> None:
        try:
            if self.person:
                self.manager.edit_person(
                    person_id=self.person.person_id,
                    full_name=self.name_edit.text(),
                    role=self.role_edit.text(),
                )
            else:
                self.manager.add_person(
                    full_name=self.name_edit.text(),
                    role=self.role_edit.text(),
                    is_active=True,
                )
            self.accept()
        except KeyManagerError as e:
            QMessageBox.warning(self, "Ошибка", str(e))


class PersonsDialog(QDialog):
    def __init__(self, manager: KeyManager, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.setWindowTitle("Получатели")
        self.resize(760, 500)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Поиск по имени...")

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["ID", "ФИО", "Роль", "Активен"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)

        btn_add = QPushButton("Добавить")
        btn_edit = QPushButton("Редактировать")
        btn_toggle = QPushButton("Отключить/Активировать")

        btn_add.clicked.connect(self._on_add)
        btn_edit.clicked.connect(self._on_edit)
        btn_toggle.clicked.connect(self._on_toggle)
        self.search_edit.textChanged.connect(self._fill)

        controls = QHBoxLayout()
        controls.addWidget(btn_add)
        controls.addWidget(btn_edit)
        controls.addWidget(btn_toggle)

        layout = QVBoxLayout(self)
        layout.addWidget(self.search_edit)
        layout.addWidget(self.table)
        layout.addLayout(controls)

        self._fill()

    def _fill(self) -> None:
        query = self.search_edit.text().strip().lower()
        persons = self.manager.get_persons(include_inactive=True)
        if query:
            persons = [p for p in persons if query in p.full_name.lower()]

        self.table.setRowCount(len(persons))
        for row, person in enumerate(persons):
            self.table.setItem(row, 0, QTableWidgetItem(str(person.person_id)))
            self.table.setItem(row, 1, QTableWidgetItem(person.full_name))
            self.table.setItem(row, 2, QTableWidgetItem(person.role))
            self.table.setItem(row, 3, QTableWidgetItem("Да" if person.is_active else "Нет"))

        self.table.resizeColumnsToContents()

    def _selected_person_id(self) -> int | None:
        row = self.table.currentRow()
        if row < 0:
            return None
        item = self.table.item(row, 0)
        if not item:
            return None
        return int(item.text())

    def _on_add(self) -> None:
        dialog = PersonEditDialog(self.manager, parent=self)
        if dialog.exec():
            self._fill()

    def _on_edit(self) -> None:
        person_id = self._selected_person_id()
        if person_id is None:
            QMessageBox.information(self, "Внимание", "Выберите человека в таблице.")
            return

        person = self.manager.get_person(person_id)
        if not person:
            return
        dialog = PersonEditDialog(self.manager, person=person, parent=self)
        if dialog.exec():
            self._fill()

    def _on_toggle(self) -> None:
        person_id = self._selected_person_id()
        if person_id is None:
            QMessageBox.information(self, "Внимание", "Выберите человека в таблице.")
            return

        person = self.manager.get_person(person_id)
        if not person:
            return

        self.manager.set_person_active(person_id, not person.is_active)
        self._fill()


class IssueDialog(QDialog):
    def __init__(self, manager: KeyManager, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.setWindowTitle("Выдача ключей")

        self.apartment_combo = QComboBox()
        self.recipient_combo = QComboBox()
        self.recipient_combo.setEditable(True)
        self.recipient_combo.setInsertPolicy(QComboBox.NoInsert)

        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 1000)

        self.limit_label = QLabel("Доступно: -")
        self.person_hint_label = QLabel("")

        form = QFormLayout()
        form.addRow("Квартира:", self.apartment_combo)
        form.addRow("Получатель:", self.recipient_combo)
        form.addRow("", self.person_hint_label)
        form.addRow("Количество:", self.count_spin)
        form.addRow("Лимит:", self.limit_label)

        self.issue_btn = QPushButton("Выдать")
        self.issue_btn.clicked.connect(self._issue)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(self.issue_btn)

        self.apartment_combo.currentIndexChanged.connect(self._update_limit)
        self._reload_data()

    def _reload_data(self) -> None:
        self.apartment_combo.clear()
        for a in self.manager.get_apartments():
            self.apartment_combo.addItem(
                f"Корпус {a.building}, этаж {a.floor}, кв. {a.apartment_number}",
                a.apartment_id,
            )

        self.recipient_combo.clear()
        active_persons = self.manager.get_persons(include_inactive=False)
        for person in active_persons:
            self.recipient_combo.addItem(f"{person.full_name} ({person.role})" if person.role else person.full_name, person.person_id)

        completer = self.recipient_combo.completer()
        if completer:
            completer.setFilterMode(Qt.MatchContains)
            completer.setCaseSensitivity(Qt.CaseInsensitive)

        if not active_persons:
            self.person_hint_label.setText("Нет активных получателей. Добавьте людей в окне 'Получатели'.")
            self.recipient_combo.setEnabled(False)
            self.issue_btn.setEnabled(False)
        else:
            self.person_hint_label.setText("Показываются только активные получатели.")
            self.recipient_combo.setEnabled(True)
            self.issue_btn.setEnabled(True)

        self._update_limit()

    def _update_limit(self) -> None:
        apartment_id = self.apartment_combo.currentData()
        apartment = self.manager.get_apartment(apartment_id) if apartment_id is not None else None
        if apartment:
            self.limit_label.setText(f"Доступно: {apartment.available_keys}")
            self.count_spin.setMaximum(max(1, apartment.available_keys))
        else:
            self.limit_label.setText("Доступно: -")

    def _issue(self) -> None:
        try:
            apartment_id = self.apartment_combo.currentData()
            if apartment_id is None:
                raise KeyManagerError("Нет доступных квартир для выдачи.")

            person_id = self.recipient_combo.currentData()
            if person_id is None:
                raise KeyManagerError("Выберите получателя из списка.")

            count = self.count_spin.value()
            self.manager.issue_keys(apartment_id=apartment_id, recipient_id=person_id, count=count)
            self.accept()
        except KeyManagerError as e:
            QMessageBox.warning(self, "Ошибка выдачи", str(e))


class ReturnDialog(QDialog):
    def __init__(self, manager: KeyManager, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.setWindowTitle("Возврат ключей")

        self.issue_combo = QComboBox()
        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 1000)
        self.remaining_label = QLabel("Осталось вернуть: -")

        form = QFormLayout()
        form.addRow("Активная выдача:", self.issue_combo)
        form.addRow("Количество возврата:", self.count_spin)
        form.addRow("Статус:", self.remaining_label)

        return_btn = QPushButton("Вернуть")
        return_btn.clicked.connect(self._return_keys)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(return_btn)

        self.issue_combo.currentIndexChanged.connect(self._update_limit)
        self._reload_data()

    def _reload_data(self) -> None:
        self.issue_combo.clear()
        for issue in self.manager.get_active_issues():
            apartment = self.manager.get_apartment(issue.apartment_id)
            if apartment:
                text = (
                    f"#{issue.issue_id} | {issue.recipient_name} | "
                    f"кв. {apartment.apartment_number} | осталось: {issue.active_count}"
                )
                self.issue_combo.addItem(text, issue.issue_id)

        self._update_limit()

    def _update_limit(self) -> None:
        issue_id = self.issue_combo.currentData()
        issue = next((x for x in self.manager.get_active_issues() if x.issue_id == issue_id), None)
        if issue:
            self.remaining_label.setText(f"Осталось вернуть: {issue.active_count}")
            self.count_spin.setMaximum(max(1, issue.active_count))
        else:
            self.remaining_label.setText("Осталось вернуть: -")

    def _return_keys(self) -> None:
        try:
            issue_id = self.issue_combo.currentData()
            if issue_id is None:
                raise KeyManagerError("Нет активных выдач для возврата.")

            self.manager.return_keys(issue_id=issue_id, count=self.count_spin.value())
            self.accept()
        except KeyManagerError as e:
            QMessageBox.warning(self, "Ошибка возврата", str(e))


class LostDialog(QDialog):
    def __init__(self, manager: KeyManager, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.setWindowTitle("Отметить утерю")

        self.apartment_combo = QComboBox()
        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 1000)
        self.limit_label = QLabel("Доступно: -")

        form = QFormLayout()
        form.addRow("Квартира:", self.apartment_combo)
        form.addRow("Количество:", self.count_spin)
        form.addRow("Лимит:", self.limit_label)

        lost_btn = QPushButton("Списать")
        lost_btn.clicked.connect(self._mark_lost)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addWidget(lost_btn)

        self.apartment_combo.currentIndexChanged.connect(self._update_limit)
        self._reload_data()

    def _reload_data(self) -> None:
        self.apartment_combo.clear()
        for a in self.manager.get_apartments():
            self.apartment_combo.addItem(
                f"Корпус {a.building}, этаж {a.floor}, кв. {a.apartment_number}",
                a.apartment_id,
            )
        self._update_limit()

    def _update_limit(self) -> None:
        apartment_id = self.apartment_combo.currentData()
        apartment = self.manager.get_apartment(apartment_id) if apartment_id is not None else None
        if apartment:
            self.limit_label.setText(f"Доступно: {apartment.available_keys}")
            self.count_spin.setMaximum(max(1, apartment.available_keys))
        else:
            self.limit_label.setText("Доступно: -")

    def _mark_lost(self) -> None:
        try:
            apartment_id = self.apartment_combo.currentData()
            if apartment_id is None:
                raise KeyManagerError("Нет доступных квартир.")

            self.manager.mark_lost(apartment_id=apartment_id, count=self.count_spin.value())
            self.accept()
        except KeyManagerError as e:
            QMessageBox.warning(self, "Ошибка", str(e))


class HistoryDialog(QDialog):
    def __init__(self, manager: KeyManager, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.setWindowTitle("История операций")
        self.resize(1100, 600)

        self.date_from_edit = QDateEdit()
        self.date_from_edit.setCalendarPopup(True)
        self.date_from_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_from_edit.setSpecialValueText("-")
        self.date_from_edit.setMinimumDate(QDate(1900, 1, 1))
        self.date_from_edit.setDate(self.date_from_edit.minimumDate())

        self.date_to_edit = QDateEdit()
        self.date_to_edit.setCalendarPopup(True)
        self.date_to_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_to_edit.setSpecialValueText("-")
        self.date_to_edit.setMinimumDate(QDate(1900, 1, 1))
        self.date_to_edit.setDate(self.date_to_edit.minimumDate())

        self.person_combo = QComboBox()
        self.person_combo.setEditable(True)
        self.person_combo.addItem("", "")
        for person in self.manager.get_persons(include_inactive=True):
            self.person_combo.addItem(person.full_name, person.full_name)
        completer = self.person_combo.completer()
        if completer:
            completer.setFilterMode(Qt.MatchContains)
            completer.setCaseSensitivity(Qt.CaseInsensitive)

        self.apartment_edit = QLineEdit()
        self.apartment_edit.setPlaceholderText("Номер квартиры")

        self.reset_btn = QPushButton("Сбросить фильтры")
        self.reset_btn.clicked.connect(self._reset_filters)

        filters = QHBoxLayout()
        filters.addWidget(QLabel("Дата с:"))
        filters.addWidget(self.date_from_edit)
        filters.addWidget(QLabel("Дата по:"))
        filters.addWidget(self.date_to_edit)
        filters.addWidget(QLabel("Человек:"))
        filters.addWidget(self.person_combo)
        filters.addWidget(QLabel("Квартира:"))
        filters.addWidget(self.apartment_edit)
        filters.addWidget(self.reset_btn)

        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(
            ["Дата и время", "Действие", "Корпус", "Этаж", "Квартира", "Получатель", "Количество", "Статус"]
        )
        self.table.setSortingEnabled(True)

        layout = QVBoxLayout(self)
        layout.addLayout(filters)
        layout.addWidget(self.table)

        self.person_combo.currentIndexChanged.connect(self._fill)
        self.person_combo.lineEdit().textChanged.connect(self._fill)
        self.apartment_edit.textChanged.connect(self._fill)
        self.date_from_edit.dateChanged.connect(self._fill)
        self.date_to_edit.dateChanged.connect(self._fill)

        self._fill()

    def _reset_filters(self) -> None:
        self.person_combo.setCurrentIndex(0)
        self.person_combo.lineEdit().clear()
        self.apartment_edit.clear()
        self.date_from_edit.setDate(self.date_from_edit.minimumDate())
        self.date_to_edit.setDate(self.date_to_edit.minimumDate())
        self._fill()

    def _fill(self) -> None:
        person_text = self.person_combo.currentText().strip().lower()
        apt_text = self.apartment_edit.text().strip().lower()

        date_from = self.date_from_edit.date()
        date_to = self.date_to_edit.date()
        has_date_from = date_from != self.date_from_edit.minimumDate()
        has_date_to = date_to != self.date_to_edit.minimumDate()

        rows = []
        for item in self.manager.get_history():
            if person_text and person_text not in item.recipient.lower():
                continue
            if apt_text and apt_text not in str(item.apartment_number).lower():
                continue

            op_date = QDate(item.timestamp.year, item.timestamp.month, item.timestamp.day)
            if has_date_from and op_date < date_from:
                continue
            if has_date_to and op_date > date_to:
                continue

            rows.append(item)

        self.table.setRowCount(len(rows))
        for row, item in enumerate(rows):
            values = [
                item.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                item.operation_type,
                item.building,
                item.floor,
                item.apartment_number,
                item.recipient,
                str(item.quantity),
                item.status,
            ]
            for col, value in enumerate(values):
                self.table.setItem(row, col, QTableWidgetItem(value))

        self.table.resizeColumnsToContents()
        self.table.sortItems(0, Qt.DescendingOrder)


class KeysOnHandDialog(QDialog):
    def __init__(self, manager: KeyManager, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.manager = manager
        self.setWindowTitle("Ключи на руках")
        self.resize(900, 500)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Корпус", "Этаж", "Квартира", "Кто держит", "Количество", "Дата выдачи"])

        layout = QVBoxLayout(self)
        layout.addWidget(self.table)

        self._fill()

    def _fill(self) -> None:
        issues = self.manager.get_active_issues()
        self.table.setRowCount(len(issues))

        for row, issue in enumerate(issues):
            apartment = self.manager.get_apartment(issue.apartment_id)
            if apartment:
                values = [
                    apartment.building,
                    str(apartment.floor),
                    apartment.apartment_number,
                    issue.recipient_name,
                    str(issue.active_count),
                    issue.issued_at.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            else:
                values = ["-", "-", "-", issue.recipient_name, str(issue.active_count), issue.issued_at.strftime("%Y-%m-%d %H:%M:%S")]

            for col, value in enumerate(values):
                self.table.setItem(row, col, QTableWidgetItem(value))

        self.table.resizeColumnsToContents()


class SettingsDialog(QDialog):
    def __init__(
        self,
        current_data_path: str,
        app_version: str,
        on_save_path: Callable[[str], bool],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._on_save_path = on_save_path
        self.setWindowTitle("Настройки")
        self.resize(650, 160)

        self.path_edit = QLineEdit(current_data_path)

        choose_btn = QPushButton("Выбрать файл")
        save_btn = QPushButton("Сохранить")
        choose_btn.clicked.connect(self._choose_file)
        save_btn.clicked.connect(self._save)

        form = QFormLayout()
        form.addRow("Версия приложения:", QLabel(app_version))
        form.addRow("Путь к data.json:", self.path_edit)

        controls = QHBoxLayout()
        controls.addWidget(choose_btn)
        controls.addWidget(save_btn)

        layout = QVBoxLayout(self)
        layout.addLayout(form)
        layout.addLayout(controls)

    def _choose_file(self) -> None:
        selected_path, _ = QFileDialog.getSaveFileName(
            self,
            "Выберите data.json",
            self.path_edit.text().strip() or "data.json",
            "JSON Files (*.json);;All Files (*)",
        )
        if selected_path:
            self.path_edit.setText(selected_path)

    def _save(self) -> None:
        path = self.path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "Ошибка", "Укажите путь к файлу data.json.")
            return
        try:
            changed = self._on_save_path(path)
            if changed:
                self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка сохранения настроек", str(e))


class MainWindow(QMainWindow):
    def __init__(
        self,
        manager: KeyManager,
        data_file_path: str,
        on_change_data_path: Callable[[str], bool],
    ) -> None:
        super().__init__()
        self.github_username = "USERNAME"
        self.github_repo = "REPO"
        self.manager = manager
        self.data_file_path = data_file_path
        self.on_change_data_path = on_change_data_path
        self.setWindowTitle("Учет ключей")
        self.resize(1100, 650)

        root = QWidget()
        self.setCentralWidget(root)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Поиск по корпусу / этажу / квартире...")

        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Корпус", "Этаж", "Квартира", "Всего ключей", "Выдано", "Утеряно", "Доступно"]
        )
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)

        btn_add = QPushButton("Добавить квартиру")
        self.btn_edit_apartment = QPushButton("Редактировать квартиру")
        btn_issue = QPushButton("Выдать ключи")
        btn_return = QPushButton("Вернуть ключи")
        btn_lost = QPushButton("Отметить утерю")
        btn_history = QPushButton("История")
        btn_persons = QPushButton("Получатели")
        btn_keys_on_hand = QPushButton("Ключи на руках")
        btn_export_excel = QPushButton("Экспорт в Excel")
        btn_check_updates = QPushButton("Проверить обновления")
        btn_settings = QPushButton("Настройки")

        btn_add.clicked.connect(self._on_add)
        self.btn_edit_apartment.clicked.connect(self._on_edit_apartment)
        btn_issue.clicked.connect(self._on_issue)
        btn_return.clicked.connect(self._on_return)
        btn_lost.clicked.connect(self._on_lost)
        btn_history.clicked.connect(self._on_history)
        btn_persons.clicked.connect(self._on_persons)
        btn_keys_on_hand.clicked.connect(self._on_keys_on_hand)
        btn_export_excel.clicked.connect(self._on_export_excel)
        btn_check_updates.clicked.connect(self._on_check_updates)
        btn_settings.clicked.connect(self._on_settings)
        self.search_edit.textChanged.connect(self.refresh_table)
        self.table.itemSelectionChanged.connect(self._update_edit_button_state)
        self.table.itemDoubleClicked.connect(self._on_apartment_double_click)

        btns = QHBoxLayout()
        for btn in (
            btn_add,
            self.btn_edit_apartment,
            btn_issue,
            btn_return,
            btn_lost,
            btn_history,
            btn_persons,
            btn_keys_on_hand,
            btn_export_excel,
            btn_check_updates,
            btn_settings,
        ):
            btns.addWidget(btn)

        layout = QVBoxLayout(root)
        layout.addWidget(self.search_edit)
        layout.addWidget(self.table)
        layout.addLayout(btns)

        self.refresh_table()
        self._update_edit_button_state()

    def refresh_table(self) -> None:
        apartments = self.manager.get_apartments()
        query = self.search_edit.text().strip().lower()

        if query:
            apartments = [
                apt
                for apt in apartments
                if query in apt.building.lower()
                or query in str(apt.floor).lower()
                or query in apt.apartment_number.lower()
            ]

        self.table.setRowCount(len(apartments))

        for row, apt in enumerate(apartments):
            values = [
                str(apt.apartment_id),
                apt.building,
                str(apt.floor),
                apt.apartment_number,
                str(apt.total_keys),
                str(apt.issued_keys),
                str(apt.lost_keys),
                str(apt.available_keys),
            ]
            for col, value in enumerate(values):
                self.table.setItem(row, col, QTableWidgetItem(value))

        self.table.resizeColumnsToContents()
        self._update_edit_button_state()

    def _selected_apartment_id(self) -> int | None:
        row = self.table.currentRow()
        if row < 0:
            return None

        item = self.table.item(row, 0)
        if not item:
            return None

        return int(item.text())

    def _update_edit_button_state(self) -> None:
        self.btn_edit_apartment.setEnabled(self._selected_apartment_id() is not None)

    def _on_add(self) -> None:
        dialog = AddApartmentDialog(self.manager, self)
        if dialog.exec():
            self.refresh_table()

    def _on_edit_apartment(self) -> None:
        apartment_id = self._selected_apartment_id()
        if apartment_id is None:
            QMessageBox.information(self, "Внимание", "Выберите квартиру в таблице.")
            return

        dialog = AddApartmentDialog(self.manager, apartment_id=apartment_id, parent=self)
        if dialog.exec():
            self.refresh_table()

    def _on_apartment_double_click(self, *_args: object) -> None:
        self._on_edit_apartment()

    def _on_issue(self) -> None:
        dialog = IssueDialog(self.manager, self)
        if dialog.exec():
            self.refresh_table()

    def _on_return(self) -> None:
        dialog = ReturnDialog(self.manager, self)
        if dialog.exec():
            self.refresh_table()

    def _on_lost(self) -> None:
        dialog = LostDialog(self.manager, self)
        if dialog.exec():
            self.refresh_table()

    def _on_history(self) -> None:
        dialog = HistoryDialog(self.manager, self)
        dialog.exec()

    def _on_persons(self) -> None:
        dialog = PersonsDialog(self.manager, self)
        dialog.exec()

    def _on_keys_on_hand(self) -> None:
        dialog = KeysOnHandDialog(self.manager, self)
        dialog.exec()

    def _on_export_excel(self) -> None:
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить Excel-файл",
            "key_manager_export.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            from openpyxl import Workbook
        except ImportError:
            QMessageBox.critical(
                self,
                "Ошибка экспорта",
                "Библиотека openpyxl не установлена. Установите зависимости из requirements.txt.",
            )
            return

        try:
            workbook = Workbook()

            apartments_sheet = workbook.active
            apartments_sheet.title = "Квартиры"
            apartments_sheet.append(["Корпус", "Этаж", "Квартира", "Всего ключей", "Выдано", "Утеряно", "Доступно"])
            for apartment in self.manager.get_apartments():
                apartments_sheet.append(
                    [
                        apartment.building,
                        apartment.floor,
                        apartment.apartment_number,
                        apartment.total_keys,
                        apartment.issued_keys,
                        apartment.lost_keys,
                        apartment.available_keys,
                    ]
                )

            history_sheet = workbook.create_sheet("История")
            history_sheet.append(["Дата/время", "Действие", "Корпус", "Этаж", "Квартира", "Получатель", "Количество", "Статус"])
            for item in self.manager.get_history():
                history_sheet.append(
                    [
                        item.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                        item.operation_type,
                        item.building,
                        item.floor,
                        item.apartment_number,
                        item.recipient,
                        item.quantity,
                        item.status,
                    ]
                )

            active_sheet = workbook.create_sheet("Ключи на руках")
            active_sheet.append(["Корпус", "Этаж", "Квартира", "Получатель", "Количество", "Дата выдачи"])
            for issue in self.manager.get_active_issues():
                apartment = self.manager.get_apartment(issue.apartment_id)
                if apartment:
                    row = [
                        apartment.building,
                        apartment.floor,
                        apartment.apartment_number,
                        issue.recipient_name,
                        issue.active_count,
                        issue.issued_at.strftime("%Y-%m-%d %H:%M:%S"),
                    ]
                else:
                    row = ["-", "-", "-", issue.recipient_name, issue.active_count, issue.issued_at.strftime("%Y-%m-%d %H:%M:%S")]
                active_sheet.append(row)

            workbook.save(file_path)
            QMessageBox.information(self, "Экспорт завершен", f"Данные успешно сохранены в файл:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", f"Не удалось сохранить файл Excel.\n\n{e}")

    def _on_settings(self) -> None:
        dialog = SettingsDialog(self.data_file_path, self._get_local_version(), self._apply_data_path, self)
        if dialog.exec():
            self.refresh_table()

    def _on_check_updates(self) -> None:
        local_version = self._get_local_version()
        try:
            remote_version = self._get_remote_version()
        except requests.RequestException as e:
            QMessageBox.critical(self, "Ошибка обновления", f"Не удалось проверить обновления.\n\nПроверьте интернет.\n{e}")
            return
        except Exception as e:
            QMessageBox.critical(self, "Ошибка обновления", f"Ошибка при проверке версии на GitHub.\n\n{e}")
            return

        if not self._is_newer_version(remote_version, local_version):
            QMessageBox.information(self, "Обновление", "У вас последняя версия.")
            return

        answer = QMessageBox.question(
            self,
            "Доступно обновление",
            (
                f"Доступна новая версия {remote_version}\n"
                f"Текущая версия: {local_version}\n\n"
                "Обновить сейчас?"
            ),
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if answer != QMessageBox.Yes:
            return

        try:
            self._download_and_apply_update()
        except requests.RequestException as e:
            QMessageBox.critical(self, "Ошибка обновления", f"Ошибка скачивания обновления.\n\n{e}")
            return
        except zipfile.BadZipFile as e:
            QMessageBox.critical(self, "Ошибка обновления", f"Ошибка распаковки архива.\n\n{e}")
            return
        except OSError as e:
            QMessageBox.critical(self, "Ошибка обновления", f"Ошибка записи файлов.\n\n{e}")
            return
        except Exception as e:
            QMessageBox.critical(self, "Ошибка обновления", f"Не удалось установить обновление.\n\n{e}")
            return

        restart = QMessageBox.question(
            self,
            "Обновление установлено",
            "Обновление установлено. Перезапустить приложение?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if restart == QMessageBox.Yes:
            self._restart_application()

    def _get_app_dir(self) -> Path:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).resolve().parent
        return Path(__file__).resolve().parent

    def _get_local_version(self) -> str:
        version_file = self._get_app_dir() / "version.txt"
        try:
            version = version_file.read_text(encoding="utf-8").strip()
            return version or "0.0.0"
        except OSError:
            return "0.0.0"

    def _get_remote_version(self) -> str:
        version_url = (
            f"https://raw.githubusercontent.com/{self.github_username}/{self.github_repo}/main/version.txt"
        )
        response = requests.get(version_url, timeout=15)
        response.raise_for_status()
        remote_version = response.text.strip()
        if not remote_version:
            raise ValueError("Пустой version.txt в репозитории.")
        return remote_version

    @staticmethod
    def _parse_version(version: str) -> tuple[int, ...]:
        parts = []
        for part in version.strip().split("."):
            digits = "".join(ch for ch in part if ch.isdigit())
            parts.append(int(digits or "0"))
        return tuple(parts)

    def _is_newer_version(self, remote: str, local: str) -> bool:
        remote_parts = self._parse_version(remote)
        local_parts = self._parse_version(local)
        max_len = max(len(remote_parts), len(local_parts))
        padded_remote = remote_parts + (0,) * (max_len - len(remote_parts))
        padded_local = local_parts + (0,) * (max_len - len(local_parts))
        return padded_remote > padded_local

    def _download_and_apply_update(self) -> None:
        zip_url = f"https://github.com/{self.github_username}/{self.github_repo}/archive/refs/heads/main.zip"

        with tempfile.TemporaryDirectory(prefix="key-manager-update-") as temp_dir:
            temp_path = Path(temp_dir)
            zip_path = temp_path / "update.zip"

            progress = QProgressDialog("Загрузка обновления...", "Отмена", 0, 100, self)
            progress.setWindowTitle("Обновление")
            progress.setWindowModality(Qt.WindowModal)
            progress.setValue(0)
            progress.show()

            with requests.get(zip_url, stream=True, timeout=30) as response:
                response.raise_for_status()
                total_size = int(response.headers.get("content-length", 0))
                downloaded = 0
                with zip_path.open("wb") as f:
                    for chunk in response.iter_content(chunk_size=65536):
                        if not chunk:
                            continue
                        f.write(chunk)
                        downloaded += len(chunk)
                        if total_size > 0:
                            progress.setValue(min(100, int((downloaded / total_size) * 100)))
                        else:
                            progress.setValue(0)
                        if progress.wasCanceled():
                            raise RuntimeError("Загрузка обновления отменена пользователем.")

            progress.setLabelText("Распаковка и установка обновления...")
            progress.setValue(100)

            extract_dir = temp_path / "unpacked"
            with zipfile.ZipFile(zip_path, "r") as archive:
                archive.extractall(extract_dir)

            source_root = self._resolve_update_source(extract_dir)
            destination_root = self._get_app_dir()
            self._sync_update_files(source_root, destination_root)

    def _resolve_update_source(self, extract_dir: Path) -> Path:
        unpacked_roots = [p for p in extract_dir.iterdir() if p.is_dir()]
        if not unpacked_roots:
            raise FileNotFoundError("Не удалось найти распакованную директорию обновления.")

        repo_root = unpacked_roots[0]
        dist_root = repo_root / "dist" / "key-manager"
        if dist_root.exists() and dist_root.is_dir():
            return dist_root
        return repo_root

    def _sync_update_files(self, source_root: Path, destination_root: Path) -> None:
        excluded_names = {"data.json", "config.json", ".git"}

        for src in source_root.rglob("*"):
            relative = src.relative_to(source_root)
            if any(part in excluded_names for part in relative.parts):
                continue

            dst = destination_root / relative
            if src.is_dir():
                dst.mkdir(parents=True, exist_ok=True)
                continue

            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)

    def _restart_application(self) -> None:
        from PySide6.QtCore import QProcess
        from PySide6.QtWidgets import QApplication

        QProcess.startDetached(sys.executable, sys.argv)
        QApplication.quit()

    def _apply_data_path(self, path: str) -> bool:
        changed = self.on_change_data_path(path)
        if changed:
            self.data_file_path = path
        return changed

    def on_data_source_changed(self, data_file_path: str) -> None:
        self.data_file_path = data_file_path
        self.refresh_table()
